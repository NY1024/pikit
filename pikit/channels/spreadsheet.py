"""Spreadsheet channel: hide the payload inside spreadsheet cell data.

Agents that analyse spreadsheets (Excel, Google Sheets, CSV-as-sheet)
read cell values, formulas, and comments. A payload planted in a cell
value, a cell comment, or a sheet name blends in with legitimate data
and is processed verbatim by the model.

**File mode**: :meth:`taint_file` operates on real ``.csv`` files.
For ``.xlsx`` files, :mod:`openpyxl` is used to inject cell values and
comments.
"""

from __future__ import annotations

import os

from ..base import Channel
from . import register


@register("spreadsheet")
class SpreadsheetChannel(Channel):
    """Embed the payload in a spreadsheet cell representation.

    Works on a plain-text spreadsheet representation (``A1: value`` lines,
    one per cell), which is what a spreadsheet-reading tool would return
    to the agent after converting a .xlsx / .gsheet to text.

    Parameters
    ----------
    method:
        * ``"cell_value"`` — appended to an existing cell value (default).
        * ``"cell_comment"`` — a cell comment (``A1 [comment]: payload``).
        * ``"sheet_name"`` — the payload becomes a sheet tab name.
    """

    def __init__(self, method: str = "cell_value") -> None:
        if method not in ("cell_value", "cell_comment", "sheet_name"):
            raise ValueError(
                "method must be 'cell_value', 'cell_comment', or 'sheet_name'"
            )
        self.method = method

    def taint(self, data: str, payload: str) -> str:
        lines = data.splitlines()
        out = []
        injected = False

        for line in lines:
            if not injected and self.method == "cell_value":
                # Append payload to the first cell value line.
                stripped = line.strip()
                if stripped and ":" in stripped:
                    ref, _, val = stripped.partition(":")
                    out.append(f"{ref}: {val} {payload}")
                    injected = True
                    continue
            out.append(line)

        if not injected:
            if self.method == "cell_value":
                out.append(f"Z99: {payload}")
            elif self.method == "cell_comment":
                one_line = " ".join(payload.splitlines())
                out.append(f"A1 [comment]: {one_line}")
            else:
                one_line = " ".join(payload.splitlines())
                out.append(f"Sheet: {one_line}")

        if self.method == "cell_comment" and not injected:
            pass  # handled above
        elif self.method == "cell_comment" and injected:
            # Also add a comment line for extra coverage.
            one_line = " ".join(payload.splitlines())
            out.append(f"A1 [comment]: {one_line}")
        elif self.method == "sheet_name" and not injected:
            pass  # handled above
        elif self.method == "sheet_name" and injected:
            one_line = " ".join(payload.splitlines())
            out.append(f"Sheet: {one_line}")

        return "\n".join(out)

    def taint_file(self, path: str, payload: str, output_path=None) -> str:
        """Inject the payload into a real spreadsheet file.

        Supports ``.csv`` files (text-based) and ``.xlsx`` files (via
        :mod:`openpyxl`).
        """
        ext = os.path.splitext(path)[1].lower()

        if output_path is None:
            base, e = os.path.splitext(path)
            output_path = f"{base}.tainted{e}"

        if ext == ".csv":
            return self._taint_csv(path, payload, output_path)
        elif ext == ".xlsx":
            return self._taint_xlsx(path, payload, output_path)
        else:
            # Fall back to text mode for unknown extensions.
            return super().taint_file(path, payload, output_path)

    def _taint_csv(self, path: str, payload: str, output_path: str) -> str:
        """Inject payload into a CSV file."""
        with open(path, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()

        if self.method == "cell_value":
            # Append payload to the last cell of the first data row.
            if len(lines) > 1:
                parts = lines[1].split(",")
                parts[-1] = f"{parts[-1]} {payload}"
                lines[1] = ",".join(parts)
            else:
                lines.append(payload)
        elif self.method == "cell_comment":
            # CSV has no native comments; prepend a comment-style line.
            one_line = " ".join(payload.splitlines())
            lines.insert(0, f"# {one_line}")
        elif self.method == "sheet_name":
            # CSV has no sheet tabs; add a sheet-name hint line.
            one_line = " ".join(payload.splitlines())
            lines.insert(0, f"# Sheet: {one_line}")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        return output_path

    def _taint_xlsx(self, path: str, payload: str, output_path: str) -> str:
        """Inject payload into an XLSX file using openpyxl."""
        from openpyxl import load_workbook
        from openpyxl.comments import Comment

        wb = load_workbook(path)

        if self.method == "sheet_name":
            one_line = " ".join(payload.splitlines())
            ws = wb.active
            ws.title = one_line[:31]  # Excel sheet name max 31 chars.
        elif self.method == "cell_comment":
            ws = wb.active
            one_line = " ".join(payload.splitlines())
            ws["A1"].comment = Comment(one_line, "reviewer")
        else:  # cell_value
            ws = wb.active
            # Append payload to the first non-empty cell in row 2.
            for row in ws.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    if cell.value is not None:
                        cell.value = f"{cell.value} {payload}"
                        break
                else:
                    continue
                break

        wb.save(output_path)
        return output_path